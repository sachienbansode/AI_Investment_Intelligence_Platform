"""The 5 BRD APIs + chat history + instruments + watchlist + agents status."""
import asyncio
from datetime import date, timedelta

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.agents.pipeline import PIPELINE_STATE, live_snapshot, run_daily_pipeline
from app.core.auth import get_current_user, require_admin
from app.core.compliance import AI_DISCLAIMER, audit_log
from app.data.aggregator import get_market_data
from app.db.database import (ChatFeedback, ChatMessage, Instrument, Portfolio,
                             SessionLocal, DeviceToken, StockScore, User,
                             UserActivity, WatchlistItem)
from app.llm.router import get_llm_router
from app.models.schemas import (AskAIRequest, AskAIResponse, PortfolioRequest,
                                PortfolioResponse, StockScoreResponse, WatchlistRequest)
from app.services import news_intel
from app.services.assistant import ask
from app.services.portfolio import analyze_portfolio

router = APIRouter(prefix="/api/v1")


# ── 1. Ask AI API ────────────────────────────────────────────────
@router.post("/ask", response_model=AskAIResponse)
async def ask_ai(req: AskAIRequest, user: User = Depends(get_current_user)):
    try:
        return await ask(req.question, req.session_id, req.language, user_id=user.id)
    except RuntimeError as e:
        # All LLM providers failed (bad/expired key, no quota, wrong model, or
        # blocked egress). Surface the reason instead of an opaque 500.
        import logging
        logging.getLogger(__name__).error("Ask AI failed: %s", e)
        raise HTTPException(
            status_code=502,
            detail=f"AI service unavailable — {e}. Check the backend logs and the "
                   f"API keys/models in backend/.env (Admin → Integrations shows which "
                   f"providers are configured).",
        )


# ── Chat history (per user) ──────────────────────────────────────
@router.get("/chat/sessions")
async def chat_sessions(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = (db.query(ChatMessage).filter_by(user_id=user.id, role="user")
                .order_by(ChatMessage.created_at.desc()).limit(500).all())
        seen, sessions = set(), []
        for r in rows:
            if r.session_id not in seen:
                seen.add(r.session_id)
                sessions.append({"session_id": r.session_id,
                                 "title": r.content[:60],
                                 "last_at": str(r.created_at)})
        return {"sessions": sessions[:10]}  # keep the last 10 conversations
    finally:
        db.close()


@router.get("/chat/history/{session_id}")
async def chat_history(session_id: str, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = (db.query(ChatMessage)
                .filter_by(user_id=user.id, session_id=session_id)
                .order_by(ChatMessage.created_at).limit(200).all())
        return {"messages": [{"role": r.role, "content": r.content,
                              "meta": r.meta or {}} for r in rows]}
    finally:
        db.close()


@router.delete("/chat/history/{session_id}")
async def delete_chat_session(session_id: str, user: User = Depends(get_current_user)):
    """Delete one conversation (all its messages) for the current user."""
    db = SessionLocal()
    try:
        n = (db.query(ChatMessage)
             .filter_by(user_id=user.id, session_id=session_id).delete())
        db.commit()
        return {"deleted": n}
    finally:
        db.close()


@router.delete("/chat/sessions")
async def clear_chat_history(user: User = Depends(get_current_user)):
    """Clear ALL chat history for the current user."""
    db = SessionLocal()
    try:
        n = db.query(ChatMessage).filter_by(user_id=user.id).delete()
        db.commit()
        return {"deleted": n}
    finally:
        db.close()


class FeedbackReq(BaseModel):
    rating: int
    session_id: str = ""
    question: str = ""
    answer: str = ""
    provider: str = ""


@router.post("/chat/feedback")
async def chat_feedback(req: FeedbackReq, user: User = Depends(get_current_user)):
    """Record a thumbs up/down on an assistant answer (assistant quality loop)."""
    if req.rating not in (1, -1):
        raise HTTPException(400, "rating must be 1 (up) or -1 (down)")
    db = SessionLocal()
    try:
        db.add(ChatFeedback(user_id=user.id, session_id=req.session_id[:120],
                            rating=req.rating, question=(req.question or "")[:2000],
                            answer=(req.answer or "")[:4000], provider=(req.provider or "")[:40]))
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.get("/chat/suggestions")
async def chat_suggestions(user: User = Depends(get_current_user)):
    """Personalised starter prompts: learned from the user's most-asked symbols
    and their watchlist, topped up with evergreen defaults."""
    from app.services.app_settings import get_setting
    label = get_setting("score_label") or "NIYTRI Score"
    db = SessionLocal()
    try:
        acts = (db.query(UserActivity)
                .filter_by(user_id=user.id, kind="symbol").all())
        liked = [a.value for a in sorted(
            acts, key=lambda r: (r.count or 0, r.last_at or 0), reverse=True)]
        watch = [w.symbol for w in
                 db.query(WatchlistItem).filter_by(user_id=user.id).limit(10).all()]
    finally:
        db.close()

    sugg = []
    for sym in liked[:2]:
        sugg.append(f"What's driving {sym}'s {label}?")
    for sym in watch[:2]:
        q = f"Latest news on {sym}"
        if q not in sugg:
            sugg.append(q)
    for d in [f"Top stocks by {label}", "What moved the market today?",
              "Summarize today's market news",
              f"Which stocks are below 50 on {label}?", "What is a P/E ratio?"]:
        if len(sugg) >= 5:
            break
        if d not in sugg:
            sugg.append(d)
    return {"suggestions": sugg[:5], "personalized": bool(liked or watch)}


class DeviceRegister(BaseModel):
    token: str
    platform: str = ""


@router.post("/devices/register")
async def register_device(req: DeviceRegister, user: User = Depends(get_current_user)):
    """Store a push-notification device token for the current user (mobile app)."""
    tok = (req.token or "").strip()
    if not tok:
        raise HTTPException(400, "token required")
    db = SessionLocal()
    try:
        row = db.query(DeviceToken).filter_by(token=tok).first()
        if row:
            row.user_id = user.id
            row.platform = req.platform or row.platform
        else:
            db.add(DeviceToken(user_id=user.id, token=tok, platform=req.platform or ""))
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.delete("/devices/{token}")
async def unregister_device(token: str, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        n = db.query(DeviceToken).filter_by(token=token, user_id=user.id).delete()
        db.commit()
        return {"deleted": n}
    finally:
        db.close()


@router.get("/indices/constituents")
async def index_constituents(user: User = Depends(get_current_user)):
    """Index -> member symbols, for filtering dashboards by index. Nifty 50 from
    the constituent seed; sectors are derived from the instruments master."""
    from app.db.database import NIFTY50_SEED
    db = SessionLocal()
    try:
        insts = db.query(Instrument).filter_by(is_active=True).all()
    finally:
        db.close()
    sectors = {}
    n50, n500, allnse = [], [], []
    for inst in insts:
        tags = inst.indices or []
        allnse.append(inst.symbol)
        if "NIFTY50" in tags:
            n50.append(inst.symbol)
        if "NIFTY500" in tags:
            n500.append(inst.symbol)
        if inst.sector:
            sectors.setdefault(inst.sector, []).append(inst.symbol)
    # Fallbacks so the filters work before a tagged re-import:
    if not n50:
        n50 = [sym for sym, _n, _s in NIFTY50_SEED]
    if not n500:
        n500 = allnse                       # current table is the Nifty 500 universe
    out = {"Nifty 50": n50, "Nifty 500": n500}
    if len(allnse) > len(n500):
        out["All NSE"] = allnse             # only show when broader than Nifty 500
    out["sectors"] = sectors
    return out


# ── Instruments (read; admin manages via /admin) ─────────────────
@router.get("/instruments")
async def instruments(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = (db.query(Instrument).filter_by(is_active=True)
                .order_by(Instrument.symbol).all())
        return {"instruments": [{"symbol": r.symbol, "name": r.name,
                                 "sector": r.sector} for r in rows]}
    finally:
        db.close()


# ── 2. Stock Score API ───────────────────────────────────────────
@router.get("/score/{symbol}", response_model=StockScoreResponse)
async def stock_score(symbol: str):
    db = SessionLocal()
    try:
        row = (db.query(StockScore).filter_by(symbol=symbol.upper())
               .order_by(StockScore.score_date.desc()).first())
    finally:
        db.close()
    if not row:
        raise HTTPException(404, f"No score for {symbol.upper()} yet.")
    return StockScoreResponse(
        symbol=row.symbol, score_date=row.score_date,
        composite_score=row.composite_score, pillar_scores=row.pillar_scores,
        explanation=row.explanation or "", quality_status=row.quality_status,
        disclaimer=AI_DISCLAIMER,
    )


@router.get("/scores")
async def all_scores(score_date: str = ""):
    from app.services.rescore import pillar_drivers
    db = SessionLocal()
    try:
        today = date.today().isoformat()
        all_dates = [d[0] for d in (db.query(StockScore.score_date).distinct()
                     .order_by(StockScore.score_date.desc()).limit(60).all())]
        latest_date = (score_date if score_date in all_dates
                       else (all_dates[0] if all_dates else today))
        # all rows for the latest run (no truncation, works for 500+ scripts)
        rows = (db.query(StockScore).filter_by(score_date=latest_date)
                .order_by(StockScore.composite_score.desc()).all())
        # Baseline for Δ Change = the most recent PRIOR run whose scores actually
        # differ from the latest. Identical re-runs (e.g. two pre-market runs that
        # read the same snapshot) are skipped so the column reflects real movement,
        # consistent with the dashboard trend deltas.
        latest_map = {r.symbol: r.composite_score for r in rows}
        prior_dates = [d[0] for d in
                       (db.query(StockScore.score_date)
                        .filter(StockScore.score_date < latest_date)
                        .distinct().order_by(StockScore.score_date.desc()).all())]
        prev_date, prev = None, {}
        for d in prior_dates:
            cand = {r.symbol: r for r in
                    db.query(StockScore).filter_by(score_date=d).all()}
            if any(round(latest_map[sym] - c.composite_score, 1) != 0
                   for sym, c in cand.items() if sym in latest_map):
                prev_date, prev = d, cand
                break
        if prev_date is None and prior_dates:   # all prior runs identical to latest
            prev_date = prior_dates[0]
            prev = {r.symbol: r for r in
                    db.query(StockScore).filter_by(score_date=prev_date).all()}
        sectors = {r.symbol: r.sector for r in db.query(Instrument).all()}

        out = []
        for r in rows:
            p = prev.get(r.symbol)
            out.append({
                "symbol": r.symbol, "composite_score": r.composite_score,
                "pillar_scores": r.pillar_scores, "explanation": r.explanation,
                "quality_status": r.quality_status,
                "sector": sectors.get(r.symbol, ""),
                "pe": r.pe, "market_cap": r.market_cap, "last_price": r.last_price,
                "fundamentals": r.fundamentals or {},
                "prev_score": p.composite_score if p else None,
                "prev_date": p.score_date if p else None,
                "delta": round(r.composite_score - p.composite_score, 1) if p else None,
                "drivers": (pillar_drivers(r.pillar_scores or {}, p.pillar_scores or {})
                            if p else []),
            })
        return {"score_date": latest_date, "dates": all_dates, "scores": out,
                "disclaimer": AI_DISCLAIMER}
    finally:
        db.close()


@router.get("/scores/{symbol}/history")
async def score_history(symbol: str, days: int = 30):
    """Daily composite score for one script over the last N days (sparkline)."""
    days = max(2, min(days, 120))
    cutoff = (date.today() - timedelta(days=days)).isoformat()
    db = SessionLocal()
    try:
        rows = (db.query(StockScore.score_date, StockScore.composite_score)
                .filter(StockScore.symbol == symbol.upper(),
                        StockScore.score_date >= cutoff)
                .order_by(StockScore.score_date).all())
        return {"symbol": symbol.upper(),
                "history": [{"date": d, "score": sc} for d, sc in rows]}
    finally:
        db.close()


@router.get("/scores/trends")
async def score_trends(days: int = 30, symbols: str = ""):
    """Daily average score + coverage for the last N days, plus the biggest
    score gainers/losers between the earliest and latest run in the window."""
    from sqlalchemy import case, func
    days = max(2, min(days, 90))
    cutoff = (date.today() - timedelta(days=days)).isoformat()
    syms = [x.strip().upper() for x in symbols.split(",") if x.strip()] if symbols else None
    db = SessionLocal()
    try:
        strong = func.sum(case((StockScore.composite_score >= 65, 1), else_=0))
        neutral = func.sum(case(((StockScore.composite_score >= 50)
                                 & (StockScore.composite_score < 65), 1), else_=0))
        weak = func.sum(case((StockScore.composite_score < 50, 1), else_=0))
        daily_rows = (db.query(StockScore.score_date,
                               func.avg(StockScore.composite_score),
                               func.count(StockScore.id),
                               func.min(StockScore.composite_score),
                               func.max(StockScore.composite_score),
                               strong, neutral, weak)
                      .filter(StockScore.score_date >= cutoff))
        if syms:
            daily_rows = daily_rows.filter(StockScore.symbol.in_(syms))
        daily_rows = (daily_rows.group_by(StockScore.score_date)
                      .order_by(StockScore.score_date).all())
        daily = [{"date": d, "avg_score": round(float(a), 1), "count": n,
                  "min_score": round(float(mn), 1), "max_score": round(float(mx), 1),
                  "strong": int(st or 0), "neutral": int(ne or 0), "weak": int(wk or 0)}
                 for d, a, n, mn, mx, st, ne, wk in daily_rows]

        gainers, losers = [], []
        if len(daily) >= 2:
            first_d, last_d = daily[0]["date"], daily[-1]["date"]
            oq = db.query(StockScore).filter_by(score_date=first_d)
            cq = db.query(StockScore).filter_by(score_date=last_d)
            if syms:
                oq = oq.filter(StockScore.symbol.in_(syms))
                cq = cq.filter(StockScore.symbol.in_(syms))
            old = {r.symbol: r.composite_score for r in oq.all()}
            cur = {r.symbol: r.composite_score for r in cq.all()}
            moves = sorted(
                ({"symbol": s, "from": old[s], "to": cur[s],
                  "delta": round(cur[s] - old[s], 1)}
                 for s in cur if s in old),
                key=lambda m: m["delta"])
            losers = [m for m in moves[:5] if m["delta"] < 0]
            gainers = [m for m in reversed(moves[-5:]) if m["delta"] > 0]
        return {"days": days, "daily": daily, "gainers": gainers, "losers": losers,
                "disclaimer": AI_DISCLAIMER}
    finally:
        db.close()


@router.post("/score/{symbol}/refresh")
async def refresh_symbol_score(symbol: str, user: User = Depends(get_current_user)):
    """Re-score one script on demand with a fresh quote; returns the change
    vs the previous score and which pillars drove it."""
    from app.services.rescore import rescore_symbol
    result = await rescore_symbol(symbol)
    if not result:
        raise HTTPException(502, f"Could not fetch a live quote for {symbol.upper()}")
    return result


# ── 3. News Summary API ──────────────────────────────────────────
@router.get("/news")
async def news_summary(limit: int = 20, refresh: bool = False):
    if refresh:
        await news_intel.refresh_news()
    items = news_intel.latest_news(limit)
    if not items:
        await news_intel.refresh_news()
        items = news_intel.latest_news(limit)
    return {"items": items, "disclaimer": AI_DISCLAIMER}


# ── 4. Portfolio Analysis API ────────────────────────────────────
@router.post("/portfolio/analyze", response_model=PortfolioResponse)
async def portfolio_analyze(req: PortfolioRequest, user: User = Depends(get_current_user)):
    if not req.holdings:
        raise HTTPException(400, "holdings cannot be empty")
    return await analyze_portfolio(req.holdings)


def _parse_portfolio_file(filename: str, data: bytes) -> list[dict]:
    """Parse an uploaded CSV/XLSX into [{symbol, quantity, avg_price}] with
    flexible, case-insensitive column matching."""
    import csv
    import io
    name = (filename or "").lower()
    records: list[dict] = []
    if name.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        grid = [[c.value for c in row] for row in wb.active.iter_rows()]
        if not grid:
            return []
        header = [str(x or "").strip().lower() for x in grid[0]]
        records = [dict(zip(header, r)) for r in grid[1:]]
    else:
        text = data.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        records = [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]

    def pick(d, keys):
        for k in keys:
            if k in d and d[k] not in (None, ""):
                return d[k]
        return None

    def num(x):
        try:
            return float(str(x).replace(",", "").strip())
        except (TypeError, ValueError):
            return 0.0

    out = []
    for d in records:
        sym = pick(d, ["symbol", "scrip", "scrip name", "scripname", "ticker", "stock",
                       "instrument", "tradingsymbol", "nse symbol", "nse"])
        qty = pick(d, ["qty", "quantity", "units", "shares", "qty."])
        price = pick(d, ["avg_price", "avg price", "average price", "avgprice",
                         "price", "buy price", "avg cost", "avg. price", "cost"])
        if sym is None and qty is None and price is None:
            continue
        out.append({"symbol": str(sym or "").strip().upper(),
                    "quantity": num(qty), "avg_price": num(price)})
    return out


@router.post("/portfolio/upload")
async def portfolio_upload(file: UploadFile = File(...),
                           user: User = Depends(get_current_user)):
    """Parse an uploaded holdings file and validate symbols against the
    instruments master. Returns a matched/unmatched summary for confirmation;
    does NOT save until the user confirms via /portfolio/save."""
    data = await file.read()
    if len(data) > 2 * 1024 * 1024:
        raise HTTPException(413, "File too large (max 2 MB).")
    if not (file.filename or "").lower().endswith((".csv", ".xlsx", ".xls", ".txt")):
        raise HTTPException(400, "Upload a .csv or .xlsx file with columns: symbol, quantity, avg_price.")
    rows = _parse_portfolio_file(file.filename, data)
    if not rows:
        raise HTTPException(422, "No holdings found. Use columns: symbol, quantity, avg_price.")
    db = SessionLocal()
    try:
        valid = {r.symbol: (r.sector or "") for r in
                 db.query(Instrument).filter_by(is_active=True).all()}
    finally:
        db.close()
    matched, unmatched = [], []
    for r in rows:
        sym = r["symbol"]
        if not sym:
            unmatched.append({**r, "reason": "missing symbol"})
        elif r["quantity"] <= 0 or r["avg_price"] <= 0:
            unmatched.append({**r, "reason": "quantity and avg price must be greater than 0"})
        elif sym not in valid:
            unmatched.append({**r, "reason": "not in the instruments master (NIFTY500) / name not matching"})
        else:
            matched.append({"symbol": sym, "quantity": r["quantity"],
                            "avg_price": r["avg_price"], "sector": valid[sym]})
    audit_log("portfolio_upload", user_id=user.id, total=len(rows),
              matched=len(matched), unmatched=len(unmatched))
    return {"matched": matched, "unmatched": unmatched,
            "counts": {"total": len(rows), "matched": len(matched),
                       "unmatched": len(unmatched)}}


@router.get("/portfolio/saved")
async def portfolio_saved(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        p = db.query(Portfolio).filter_by(user_id=user.id).first()
        return {"holdings": (p.holdings or []) if p else [],
                "updated_at": str(p.updated_at) if p else None}
    finally:
        db.close()


@router.post("/portfolio/save")
async def portfolio_save(req: PortfolioRequest, user: User = Depends(get_current_user)):
    """Persist the user's holdings (kept across sessions)."""
    holdings = [h.model_dump() for h in req.holdings]
    db = SessionLocal()
    try:
        p = db.query(Portfolio).filter_by(user_id=user.id).first()
        if p:
            p.holdings = holdings
        else:
            db.add(Portfolio(user_id=user.id, holdings=holdings))
        db.commit()
    finally:
        db.close()
    audit_log("portfolio_saved", user_id=user.id, holdings=len(holdings))
    return {"saved": len(holdings)}


@router.post("/portfolio/report.pdf")
async def portfolio_report_pdf(req: PortfolioRequest, user: User = Depends(get_current_user)):
    """Generate a shareable PDF of the portfolio analysis."""
    if not req.holdings:
        raise HTTPException(400, "holdings cannot be empty")
    analysis = await analyze_portfolio(req.holdings)
    from app.services.portfolio_pdf import build_portfolio_pdf
    holdings = [h.model_dump() for h in req.holdings]
    pdf = build_portfolio_pdf(analysis.model_dump(), holdings)
    audit_log("portfolio_pdf", user_id=user.id, holdings=len(holdings))
    import io
    return StreamingResponse(
        io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="portfolio_analysis.pdf"'})


@router.get("/compare")
async def compare(a: str, b: str, language: str = "en",
                  user: User = Depends(get_current_user)):
    """Side-by-side comparison of two NSE scripts with an advice-free AI summary."""
    a2, b2 = (a or "").upper().strip(), (b or "").upper().strip()
    if not a2 or not b2:
        raise HTTPException(400, "Provide two symbols")
    if a2 == b2:
        raise HTTPException(400, "Choose two different symbols")
    from app.services.assistant import compare_stocks
    return await compare_stocks(a2, b2, language)


@router.get("/compare/random")
async def compare_random(language: str = "en", user: User = Depends(get_current_user)):
    """Pick two random active instruments from the SAME sector and compare them."""
    import random
    from collections import defaultdict
    db = SessionLocal()
    try:
        rows = db.query(Instrument).filter_by(is_active=True).all()
        by_sector = defaultdict(list)
        for i in rows:
            if i.sector:
                by_sector[i.sector].append(i.symbol)
    finally:
        db.close()
    pools = [syms for syms in by_sector.values() if len(syms) >= 2]
    if not pools:
        raise HTTPException(400, "Not enough instruments in any single sector to compare")
    a, b = random.sample(random.choice(pools), 2)
    from app.services.assistant import compare_stocks
    return await compare_stocks(a, b, language)


@router.get("/portfolio/template.csv")
async def portfolio_template(user: User = Depends(get_current_user)):
    """Downloadable CSV pre-filled with every active script (NIFTY500) and its
    current LTP in the avg_price column. The user edits quantity/price and
    re-uploads. LTP is best-effort within a time budget; blanks where a live
    quote couldn't be fetched in time."""
    import csv
    import io
    db = SessionLocal()
    try:
        insts = [(r.symbol, r.name or "", r.sector or "") for r in
                 db.query(Instrument).filter_by(is_active=True)
                 .order_by(Instrument.symbol).all()]
    finally:
        db.close()
    md = get_market_data()
    sem = asyncio.Semaphore(20)
    ltp: dict[str, float] = {}

    async def fetch(sym):
        async with sem:
            try:
                q = await asyncio.wait_for(md.get_quote(sym), timeout=3)
                if q and q.last_price:
                    ltp[sym] = q.last_price
            except Exception:
                pass

    try:
        await asyncio.wait_for(asyncio.gather(*(fetch(s) for s, _, _ in insts)), timeout=25)
    except asyncio.TimeoutError:
        pass  # use whatever LTPs we collected; the rest stay blank

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["symbol", "name", "sector", "quantity", "avg_price"])
    for sym, name, sec in insts:
        w.writerow([sym, name, sec, "", ltp.get(sym, "")])
    audit_log("portfolio_template", user_id=user.id, scripts=len(insts), priced=len(ltp))
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")), media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="portfolio_template.csv"'})


# ── 5. Watchlist (persistent, per user) ──────────────────────────
async def _watchlist_rows(symbols: list[str]) -> list[dict]:
    md = get_market_data()
    quotes = await asyncio.gather(*(md.get_quote(s) for s in symbols))
    db = SessionLocal()
    try:
        out = []
        for sym, q in zip(symbols, quotes):
            row = (db.query(StockScore).filter_by(symbol=sym.upper())
                   .order_by(StockScore.score_date.desc()).first())
            out.append({
                "symbol": sym.upper(),
                "last_price": q.last_price if q else None,
                "change_pct": q.change_pct if q else None,
                "ai_score": row.composite_score if row else None,
                "score_date": row.score_date if row else None,
            })
        return out
    finally:
        db.close()


@router.get("/watchlist")
async def get_watchlist(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        symbols = [r.symbol for r in
                   db.query(WatchlistItem).filter_by(user_id=user.id)
                   .order_by(WatchlistItem.created_at).all()]
    finally:
        db.close()
    return {"watchlist": await _watchlist_rows(symbols), "disclaimer": AI_DISCLAIMER}


@router.post("/watchlist/insights")
async def watchlist_insights(req: WatchlistRequest, user: User = Depends(get_current_user)):
    """BRD Watchlist Insights API — ad-hoc symbol list."""
    return {"watchlist": await _watchlist_rows(req.symbols), "disclaimer": AI_DISCLAIMER}


@router.post("/watchlist/{symbol}")
async def add_to_watchlist(symbol: str, user: User = Depends(get_current_user)):
    symbol = symbol.upper()
    db = SessionLocal()
    try:
        if not db.query(Instrument).filter_by(symbol=symbol, is_active=True).first():
            raise HTTPException(404, f"Unknown instrument {symbol}")
        if db.query(WatchlistItem).filter_by(user_id=user.id, symbol=symbol).first():
            raise HTTPException(409, f"{symbol} already in watchlist")
        db.add(WatchlistItem(user_id=user.id, symbol=symbol))
        db.commit()
    finally:
        db.close()
    return {"added": symbol}


@router.delete("/watchlist/{symbol}")
async def remove_from_watchlist(symbol: str, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        n = (db.query(WatchlistItem)
             .filter_by(user_id=user.id, symbol=symbol.upper()).delete())
        db.commit()
    finally:
        db.close()
    if not n:
        raise HTTPException(404, "Not in watchlist")
    return {"removed": symbol.upper()}


# ── Agents dashboard ─────────────────────────────────────────────
@router.get("/agents/status")
async def agents_status(user: User = Depends(get_current_user)):
    from app.main import scheduler  # late import avoids circular dependency
    from app.services.app_settings import get_setting
    freq = {
        "daily_scoring": f"daily at {int(get_setting('daily_scoring_hour')):02d}:00 (server time)",
        "news_refresh": f"every {int(get_setting('news_refresh_minutes'))} minutes",
    }
    jobs = [{"id": j.id,
             "frequency": freq.get(j.id, ""),
             "next_run": j.next_run_time.isoformat() if j.next_run_time else None}
            for j in scheduler.get_jobs()]
    running = live_snapshot()
    return {
        "running": bool(running),
        "active_agents": ([a["name"] for a in running["agents"]
                           if a["status"] == "running"] if running else []),
        "current": running,
        "last": PIPELINE_STATE["last"],
        "history": PIPELINE_STATE["history"],
        "scheduled_jobs": jobs,
        "llm_providers": get_llm_router().active_providers,
        "market_data_providers": get_market_data().active_providers,
    }


# ── Market snapshot + ops ────────────────────────────────────────
@router.get("/market/indices")
async def market_indices():
    return {"indices": await get_market_data().get_indices()}


@router.post("/admin/run-scoring", dependencies=[Depends(require_admin)])
async def trigger_scoring(background: BackgroundTasks, full: bool = False):
    if PIPELINE_STATE["current"]:
        raise HTTPException(409, "Pipeline already running")
    from app.services.app_settings import get_setting
    incremental = (not full) and bool(get_setting("incremental_rescore_enabled"))
    background.add_task(run_daily_pipeline, None, incremental)
    return {"status": "scoring pipeline started",
            "mode": "incremental (only missing/failed today)" if incremental else "full re-score"}


@router.get("/branding")
async def branding():
    """Public: the admin-uploaded logo (data URI) or empty. Used for the app
    logo + favicon, including on the pre-login screen."""
    from app.services.app_settings import get_setting
    return {"logo": get_setting("brand_logo") or "",
            "score_label": get_setting("score_label") or "NIYTRI Score",
            "platform_label": get_setting("platform_label") or "NIYTRI AI",
            "ticker_position": get_setting("ticker_position") or "top"}


@router.get("/health")
async def health():
    return {
        "status": "ok",
        "llm_providers": get_llm_router().active_providers,
        "market_data_providers": get_market_data().active_providers,
    }
