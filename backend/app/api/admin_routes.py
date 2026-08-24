"""Admin APIs (admin role only): audit log browser, LLM usage stats,
score quality review (maker-checker), user management."""
import calendar
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field

from app.core.auth import hash_password, require_admin
from app.core.compliance import audit_log
from app.db.database import (ALL_PAGES, ChatFeedback, Instrument, InviteCode, PartnerKey,
                             PipelineRun, Role, SessionLocal, StockScore, User, Waitlist,
                             Invitation, EmailLog, TosAcceptance, TermsVersion)
from app.services.app_settings import DEFAULTS, all_settings, get_setting, set_setting

IST = ZoneInfo("Asia/Kolkata")


def fmt_ist(dt) -> str:
    """DDMMMYYYY hh:mm:ss AM/PM in IST."""
    if dt is None:
        return ""
    if isinstance(dt, (int, float)):
        dt = datetime.fromtimestamp(dt, tz=timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST).strftime("%d%b%Y %I:%M:%S %p")

router = APIRouter(prefix="/api/v1/admin", tags=["admin"],
                   dependencies=[Depends(require_admin)])

def _read_audit(limit: int = 5000) -> list[dict]:
    from app.core.compliance import audit_log_path
    audit_file = Path(audit_log_path())
    if not audit_file.exists():
        return []
    lines = audit_file.read_text(encoding="utf-8", errors="ignore").splitlines()
    records = []
    for line in lines[-limit:]:
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return records


@router.get("/audit")
def audit_browser(event: str = "", limit: int = 100, offset: int = 0):
    records = [r for r in reversed(_read_audit())
               if not event or r.get("event") == event]
    total = len(records)
    page = records[offset:offset + limit]
    for r in page:
        r["time"] = datetime.fromtimestamp(r["ts"], tz=timezone.utc).isoformat()
    events = sorted({r.get("event", "") for r in records})
    return {"total": total, "events": events, "records": page}


@router.get("/stats")
def usage_stats():
    records = _read_audit()
    llm_calls = [r for r in records if r.get("event") == "llm_call"]
    by_provider: dict[str, int] = {}
    by_task: dict[str, int] = {}
    tokens_in = tokens_out = 0
    for r in llm_calls:
        by_provider[r.get("provider", "?")] = by_provider.get(r.get("provider", "?"), 0) + 1
        by_task[r.get("task", "?")] = by_task.get(r.get("task", "?"), 0) + 1
        usage = r.get("usage") or {}
        tokens_in += usage.get("input_tokens", 0) or 0
        tokens_out += usage.get("output_tokens", 0) or 0
    pipelines = [r for r in records if r.get("event") == "pipeline_complete"]
    logins = len([r for r in records if r.get("event") == "login_success"])
    db = SessionLocal()
    try:
        n_users = db.query(User).count()
        n_scores = db.query(StockScore).count()
    finally:
        db.close()
    return {
        "llm_calls_total": len(llm_calls), "llm_calls_by_provider": by_provider,
        "llm_calls_by_task": by_task,
        "tokens": {"input": tokens_in, "output": tokens_out},
        "pipeline_runs": len(pipelines),
        "last_pipeline": pipelines[-1] if pipelines else None,
        "logins": logins, "users": n_users, "scores_stored": n_scores,
    }


class ReviewRequest(BaseModel):
    status: str  # approved | rejected


@router.get("/scores/pending")
def pending_scores():
    db = SessionLocal()
    try:
        rows = (db.query(StockScore).order_by(StockScore.score_date.desc(),
                                              StockScore.symbol).limit(200).all())
        return [{
            "id": r.id, "symbol": r.symbol, "score_date": r.score_date,
            "composite_score": r.composite_score, "quality_status": r.quality_status,
            "explanation": r.explanation, "reviewed_by": r.reviewed_by,
            "ai_review": r.ai_review,
        } for r in rows]
    finally:
        db.close()


@router.get("/scores/history")
def scores_history(score_date: str = "", status: str = "", symbol: str = "",
                   limit: int = 100, offset: int = 0):
    """Full audit of scores across all runs: filterable, with per-date
    approved/rejected summary and reviewer attribution (auto vs human)."""
    from sqlalchemy import func
    db = SessionLocal()
    try:
        q = db.query(StockScore)
        if score_date:
            q = q.filter(StockScore.score_date == score_date)
        if status:
            q = q.filter(StockScore.quality_status == status)
        if symbol:
            q = q.filter(StockScore.symbol.like(f"%{symbol.upper()}%"))
        total = q.count()
        rows = (q.order_by(StockScore.score_date.desc(), StockScore.symbol)
                .offset(offset).limit(limit).all())

        summary_q = (db.query(StockScore.score_date, StockScore.quality_status,
                              func.count(StockScore.id))
                     .group_by(StockScore.score_date, StockScore.quality_status)
                     .order_by(StockScore.score_date.desc()).all())
        summary: dict[str, dict] = {}
        for d, st, n in summary_q:
            summary.setdefault(d, {"score_date": d, "approved": 0,
                                   "rejected": 0, "pending": 0})
            summary[d][st or "pending"] = n
        human = (db.query(func.count(StockScore.id))
                 .filter(StockScore.reviewed_by != "").scalar() or 0)

        return {
            "total": total,
            "human_reviewed_total": human,
            "summary": list(summary.values())[:30],
            "rows": [{
                "id": r.id, "symbol": r.symbol, "score_date": r.score_date,
                "composite_score": r.composite_score,
                "quality_status": r.quality_status,
                "explanation": r.explanation,
                "pillar_scores": r.pillar_scores,
                "ai_review": r.ai_review,
                "reviewed_by": r.reviewed_by or "auto (Quality Agent)",
                "reviewed_at": str(r.reviewed_at) if r.reviewed_at else None,
                "created_at": str(r.created_at),
            } for r in rows],
        }
    finally:
        db.close()


@router.patch("/scores/{score_id}/review")
def review_score(score_id: int, req: ReviewRequest,
                 admin: User = Depends(require_admin)):
    if req.status not in ("approved", "rejected"):
        raise HTTPException(400, "status must be 'approved' or 'rejected'")
    db = SessionLocal()
    try:
        row = db.get(StockScore, score_id)
        if not row:
            raise HTTPException(404, "Score not found")
        row.quality_status = req.status
        row.reviewed_by = admin.email
        row.reviewed_at = datetime.now(timezone.utc)
        db.commit()
        audit_log("score_review", score_id=score_id, symbol=row.symbol,
                  status=req.status, reviewer=admin.email)
        return {"id": row.id, "symbol": row.symbol, "quality_status": row.quality_status}
    finally:
        db.close()


class BulkReviewRequest(BaseModel):
    set_status: str            # approved | rejected
    score_date: str = ""       # limit to this run date
    status: str = ""           # only rows currently in this status (e.g. pending)
    symbol: str = ""           # symbol contains


@router.post("/scores/review-bulk")
def review_scores_bulk(req: BulkReviewRequest, admin: User = Depends(require_admin)):
    """Approve or reject every score matching the given filters in one action.
    Filters mirror the Score-review screen (run date, current status, symbol)."""
    if req.set_status not in ("approved", "rejected"):
        raise HTTPException(400, "set_status must be 'approved' or 'rejected'")
    db = SessionLocal()
    try:
        q = db.query(StockScore)
        if req.score_date:
            q = q.filter(StockScore.score_date == req.score_date)
        if req.status:
            q = q.filter(StockScore.quality_status == req.status)
        if req.symbol:
            q = q.filter(StockScore.symbol.like(f"%{req.symbol.upper()}%"))
        n = q.update({StockScore.quality_status: req.set_status,
                      StockScore.reviewed_by: admin.email,
                      StockScore.reviewed_at: datetime.now(timezone.utc)},
                     synchronize_session=False)
        db.commit()
        audit_log("score_review_bulk", set_status=req.set_status, count=n,
                  score_date=req.score_date or "all", status_filter=req.status or "all",
                  symbol=req.symbol or "", reviewer=admin.email)
        return {"updated": n}
    finally:
        db.close()


@router.get("/users")
def list_users():
    db = SessionLocal()
    try:
        return [{"id": u.id, "email": u.email, "full_name": u.full_name,
                 "is_admin": bool(u.is_admin), "is_active": bool(u.is_active),
                 "role_id": u.role_id, "created_at": str(u.created_at)}
                for u in db.query(User).all()]
    finally:
        db.close()


class CreateUserRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=8, max_length=128)
    full_name: str = ""
    is_admin: bool = False
    role_id: int | None = None


@router.post("/users")
def create_user(req: CreateUserRequest, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        if db.query(User).filter_by(email=req.email.lower()).first():
            raise HTTPException(409, "Email already exists")
        user = User(email=req.email.lower(), full_name=req.full_name,
                    hashed_password=hash_password(req.password), is_admin=req.is_admin,
                    role_id=req.role_id)
        db.add(user)
        db.commit()
        db.refresh(user)
    finally:
        db.close()
    audit_log("user_created", user=user.email, admin=req.is_admin, by=admin.email)
    return {"id": user.id, "email": user.email, "is_admin": bool(user.is_admin)}


@router.patch("/users/{user_id}/toggle-active")
def toggle_user(user_id: int, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        user = db.get(User, user_id)
        if not user:
            raise HTTPException(404, "User not found")
        if user.id == admin.id:
            raise HTTPException(400, "Cannot disable your own account")
        user.is_active = not user.is_active
        db.commit()
        result = {"id": user.id, "email": user.email, "is_active": bool(user.is_active)}
    finally:
        db.close()
    audit_log("user_toggled", **result, by=admin.email)
    return result


# ── Instruments management ───────────────────────────────────────
class InstrumentRequest(BaseModel):
    symbol: str = Field(min_length=1, max_length=20)
    name: str = ""
    sector: str = ""


@router.get("/instruments")
def admin_instruments():
    db = SessionLocal()
    try:
        rows = db.query(Instrument).order_by(Instrument.symbol).all()
        return [{"id": r.id, "symbol": r.symbol, "name": r.name, "sector": r.sector,
                 "is_active": bool(r.is_active),
                 "in_scoring_universe": bool(r.in_scoring_universe)} for r in rows]
    finally:
        db.close()


@router.post("/instruments")
def add_instrument(req: InstrumentRequest, admin: User = Depends(require_admin)):
    symbol = req.symbol.strip().upper()
    db = SessionLocal()
    try:
        if db.query(Instrument).filter_by(symbol=symbol).first():
            raise HTTPException(409, f"{symbol} already exists")
        row = Instrument(symbol=symbol, name=req.name.strip(), sector=req.sector.strip())
        db.add(row)
        db.commit()
        db.refresh(row)
    finally:
        db.close()
    audit_log("instrument_added", symbol=symbol, by=admin.email)
    return {"id": row.id, "symbol": row.symbol}


@router.patch("/instruments/{inst_id}/toggle/{field}")
def toggle_instrument(inst_id: int, field: str, admin: User = Depends(require_admin)):
    if field not in ("is_active", "in_scoring_universe"):
        raise HTTPException(400, "field must be is_active or in_scoring_universe")
    db = SessionLocal()
    try:
        row = db.get(Instrument, inst_id)
        if not row:
            raise HTTPException(404, "Instrument not found")
        setattr(row, field, not getattr(row, field))
        db.commit()
        result = {"id": row.id, "symbol": row.symbol, field: bool(getattr(row, field))}
    finally:
        db.close()
    audit_log("instrument_toggled", **result, by=admin.email)
    return result


async def _import_nse_csv(url: str, tag: str, include_in_scoring: bool):
    """Download an NSE official CSV and upsert instruments, tagging index
    membership. Handles the index-constituent CSVs (Symbol/Company Name/Industry)
    and the full equity master EQUITY_L.csv (SYMBOL/NAME OF COMPANY)."""
    import csv
    import io

    import httpx

    headers = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
               "Referer": "https://www.nseindia.com/"}
    try:
        async with httpx.AsyncClient(timeout=45, follow_redirects=True) as client:
            r = await client.get(url, headers=headers)
            r.raise_for_status()
    except Exception as e:
        raise HTTPException(502, f"Could not download list from NSE ({tag}): {e}")

    reader = csv.DictReader(io.StringIO(r.text))
    added = updated = 0
    db = SessionLocal()
    try:
        for row in reader:
            symbol = (row.get("Symbol") or row.get("SYMBOL") or "").strip().upper()
            if not symbol:
                continue
            series = (row.get(" SERIES") or row.get("SERIES") or "EQ").strip()
            if series and series not in ("EQ", "BE", ""):
                continue
            name = (row.get("Company Name") or row.get("NAME OF COMPANY") or "").strip()
            sector = (row.get("Industry") or "").strip()
            inst = db.query(Instrument).filter_by(symbol=symbol).first()
            if inst:
                inst.name = name or inst.name
                inst.sector = sector or inst.sector
                tags = set(inst.indices or [])
                tags.add(tag)
                inst.indices = sorted(tags)
                updated += 1
            else:
                db.add(Instrument(symbol=symbol, name=name, sector=sector,
                                  in_scoring_universe=include_in_scoring, indices=[tag]))
                added += 1
        db.commit()
        total = db.query(Instrument).count()
    finally:
        db.close()
    audit_log("instruments_import", index=tag, added=added, updated=updated)
    return {"added": added, "updated": updated, "total_instruments": total, "index": tag,
            "note": "New scripts are " + ("included in" if include_in_scoring else "excluded from")
                    + " daily scoring. Adjust per script in Admin -> Instruments."}


@router.post("/instruments/import-nifty50")
async def import_nifty50(admin: User = Depends(require_admin)):
    """Import/refresh the NIFTY 50 constituents (tagged NIFTY50, scored daily)."""
    return await _import_nse_csv(
        "https://nsearchives.nseindia.com/content/indices/ind_nifty50list.csv", "NIFTY50", True)


@router.post("/instruments/import-nifty500")
async def import_nifty500(include_in_scoring: bool = True,
                          admin: User = Depends(require_admin)):
    """Import/refresh the NIFTY 500 universe (tagged NIFTY500, scored daily)."""
    return await _import_nse_csv(
        "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv",
        "NIFTY500", include_in_scoring)


@router.post("/instruments/import-nse-all")
async def import_nse_all(admin: User = Depends(require_admin)):
    """Import the FULL NSE equity master (EQUITY_L.csv, tagged NSE). Added OUTSIDE
    the daily scoring universe (scored on-demand) to control cost."""
    return await _import_nse_csv(
        "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv", "NSE", False)


# ── Sector backfill (results persisted to instruments.sector in the DB) ──────
@router.get("/instruments/sectors-status")
def sectors_status(admin: User = Depends(require_admin)):
    """How many active scripts have / lack a sector."""
    db = SessionLocal()
    try:
        active = db.query(Instrument).filter_by(is_active=True).all()
    finally:
        db.close()
    blank = sum(1 for r in active if not (r.sector or "").strip())
    return {"active": len(active), "with_sector": len(active) - blank, "blank": blank}


@router.post("/instruments/backfill-sectors")
async def backfill_sectors(source: str = "nse", limit: int = 150,
                           overwrite: bool = False,
                           admin: User = Depends(require_admin)):
    """Fill blank instrument sectors and store them in the DB.
    source='nse'   → NSE broad-universe classification CSVs (Total Market /
                     Microcap / Smallcap / Midcap / NIFTY500), covers the long tail.
    source='yahoo' → Yahoo assetProfile per script (capped by `limit`)."""
    from app.services import sector_map as sm
    src = (source or "nse").lower()
    if src == "yahoo":
        from app.data.aggregator import get_market_data
        blanks = [r["symbol"] for r in sm.blank_sector_symbols()][:max(1, min(limit, 500))]
        if not blanks:
            return {"source": "yahoo", "matched": 0, "updated": 0, "blank_after": 0,
                    "note": "No blank sectors to fill."}
        mapping = await get_market_data().get_sectors(blanks)
        res = sm.apply_sector_map(mapping, overwrite=overwrite)
        res.update(source="yahoo", attempted=len(blanks))
    else:
        mapping = await sm.fetch_nse_sector_map()
        if not mapping:
            raise HTTPException(502, "Could not download any NSE classification list "
                                     "(network blocked or all URLs failed).")
        res = sm.apply_sector_map(mapping, overwrite=overwrite)
        res.update(source="nse")
    audit_log("instruments_sector_backfill", by=admin.email, **{k: res[k] for k in
              ("source", "matched", "updated", "blank_after") if k in res})
    res["note"] = ("Sectors stored in the DB. Click again to fill more"
                   if res.get("blank_after") else "All sectors filled.")
    return res


@router.post("/instruments/sector-map-upload")
async def sector_map_upload(file: UploadFile = File(...), overwrite: bool = Form(False),
                            admin: User = Depends(require_admin)):
    """Upload a one-time mapping CSV (symbol/ISIN + sector/industry); fills blank
    instrument sectors and stores them in the DB."""
    from app.services import sector_map as sm
    if not (file.filename or "").lower().endswith((".csv", ".txt")):
        raise HTTPException(400, "Please upload a .csv mapping file.")
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(413, "Mapping file too large (max 5 MB).")
    try:
        text = raw.decode("utf-8-sig", errors="ignore")
        mapping = sm.parse_sector_csv(text)
    except Exception as e:
        raise HTTPException(422, f"Could not parse mapping CSV: {e}")
    if not mapping:
        raise HTTPException(422, "No symbol→sector rows found. Expected columns like "
                                 "'Symbol' and 'Industry'/'Sector'.")
    res = sm.apply_sector_map(mapping, overwrite=bool(overwrite))
    res["source"] = "upload"
    audit_log("instruments_sector_map_upload", by=admin.email,
              rows=res.get("map_size"), matched=res.get("matched"),
              updated=res.get("updated"), filename=file.filename)
    return res


# ── Pipeline run audit ───────────────────────────────────────────
@router.get("/pipeline-runs")
def pipeline_runs(search: str = "", status: str = "", limit: int = 20, offset: int = 0):
    db = SessionLocal()
    try:
        q = db.query(PipelineRun)
        if search:
            q = q.filter(PipelineRun.run_id.like(f"%{search}%"))
        if status:
            q = q.filter(PipelineRun.status == status)
        total = q.count()
        rows = (q.order_by(PipelineRun.started.desc())
                .offset(offset).limit(limit).all())
        return {"total": total, "rows": [{
            "run_id": r.run_id,
            "started_ist": fmt_ist(r.started), "finished_ist": fmt_ist(r.finished),
            "duration_s": round((r.finished - r.started).total_seconds(), 1)
                          if r.started and r.finished else None,
            "status": r.status, "symbols_count": r.symbols_count,
            "agents": [{**a,
                        "started_ist": fmt_ist(a.get("started")),
                        "finished_ist": fmt_ist(a.get("finished"))}
                       for a in (r.agents or [])],
        } for r in rows]}
    finally:
        db.close()


@router.get("/pipeline-runs/export")
def export_pipeline_runs(search: str = "", status: str = ""):
    """Download the full run audit as an Excel workbook (Runs + Agent details)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = SessionLocal()
    try:
        q = db.query(PipelineRun)
        if search:
            q = q.filter(PipelineRun.run_id.like(f"%{search}%"))
        if status:
            q = q.filter(PipelineRun.status == status)
        rows = q.order_by(PipelineRun.started.desc()).limit(5000).all()
    finally:
        db.close()

    if not rows:
        raise HTTPException(404, "No pipeline runs match the current filter — "
                                 "nothing to export. Run the scoring pipeline first.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline Runs"
    headers = ["Run ID", "Started (IST)", "Finished (IST)", "Duration (s)",
               "Status", "Scripts", "Symbols"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.run_id, fmt_ist(r.started), fmt_ist(r.finished),
                   round((r.finished - r.started).total_seconds(), 1)
                   if r.started and r.finished else "",
                   r.status, r.symbols_count,
                   ", ".join(r.symbols or [])[:1000]])

    ws2 = wb.create_sheet("Agent Details")
    ws2.append(["Run ID", "Agent", "Status", "Started (IST)", "Finished (IST)",
                "Duration (s)", "Detail"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in rows:
        for a in (r.agents or []):
            dur = (round(a["finished"] - a["started"], 1)
                   if a.get("started") and a.get("finished") else "")
            ws2.append([r.run_id, a.get("name"), a.get("status"),
                        fmt_ist(a.get("started")), fmt_ist(a.get("finished")),
                        dur, a.get("detail", "")])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            width = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pipeline_runs_{datetime.now(IST).strftime('%d%b%Y_%I%M%S%p')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── LLM utilization & INR billing ────────────────────────────────
@router.get("/llm-usage")
def llm_usage():
    """Token utilization and INR cost from the audit trail: per provider,
    per pipeline stage, today / MTD actuals and month-end estimate."""
    pricing = get_setting("llm_pricing")
    usd_inr = float(pricing.get("usd_inr", 94.5))
    now = datetime.now(IST)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    days_in_month = calendar.monthrange(now.year, now.month)[1]

    def cost_inr(provider: str, tin: int, tout: int) -> float:
        p = pricing.get(provider, {})
        usd = (tin / 1e6) * float(p.get("input_usd_per_mtok", 0)) \
            + (tout / 1e6) * float(p.get("output_usd_per_mtok", 0))
        return usd * usd_inr

    by_provider: dict = {}
    by_task: dict = {}
    mtd_calls = 0
    today_cost = 0.0
    for rec in _read_audit(50000):
        if rec.get("event") != "llm_call":
            continue
        ts = datetime.fromtimestamp(rec["ts"], tz=timezone.utc).astimezone(IST)
        if ts < month_start:
            continue
        prov = rec.get("provider", "?")
        task = rec.get("task", "?")
        usage = rec.get("usage") or {}
        tin = usage.get("input_tokens") or usage.get("prompt_tokens") or 0
        tout = usage.get("output_tokens") or usage.get("completion_tokens") or 0
        c = cost_inr(prov, tin, tout)
        mtd_calls += 1
        p = by_provider.setdefault(prov, {"calls": 0, "input_tokens": 0,
                                          "output_tokens": 0, "cost_inr": 0.0})
        p["calls"] += 1; p["input_tokens"] += tin
        p["output_tokens"] += tout; p["cost_inr"] += c
        t = by_task.setdefault(task, {"calls": 0, "input_tokens": 0,
                                      "output_tokens": 0, "cost_inr": 0.0})
        t["calls"] += 1; t["input_tokens"] += tin
        t["output_tokens"] += tout; t["cost_inr"] += c
        if ts.date() == now.date():
            today_cost += c

    mtd_cost = sum(p["cost_inr"] for p in by_provider.values())
    est_month = mtd_cost / max(now.day, 1) * days_in_month
    for d in (*by_provider.values(), *by_task.values()):
        d["cost_inr"] = round(d["cost_inr"], 2)

    return {
        "as_of_ist": fmt_ist(datetime.now(timezone.utc)),
        "month": now.strftime("%b %Y"),
        "mtd": {"calls": mtd_calls, "cost_inr": round(mtd_cost, 2),
                "today_cost_inr": round(today_cost, 2)},
        "month_estimate_inr": round(est_month, 2),
        "by_provider": by_provider,
        "by_stage": by_task,
        "pricing": pricing,
        "note": "Estimates from audit-log token counts and the configurable rates in "
                "Settings (llm_pricing). Verify against provider invoices. Gemini "
                "token counts are not captured and bill as 0 here.",
    }


@router.get("/integrations")
def integrations():
    """All data/AI integrations. Public endpoints are shown in full (no secrets
    involved); the firm's own API keys are masked — full keys never leave the
    server (.env)."""
    from app.config import get_settings
    from app.data.rss_news import FEEDS
    from app.services.app_settings import get_setting
    s = get_settings()
    llm_en = get_setting("llm_enabled") or {}
    src_en = get_setting("market_sources_enabled") or {}

    def mask(key: str) -> str:
        if not key:
            return ""
        if len(key) > 12:
            return key[:4] + "•" * 10 + key[-4:]
        return "•" * 12

    from app.services.app_settings import llm_key as _llm_key, llm_base as _llm_base
    _models = get_setting("llm_models") or {}
    _dbkeys = get_setting("llm_api_keys") or {}

    def _prov(name, key, default_model, endpoint):
        resolved = _llm_key(key)
        src = "db" if (_dbkeys.get(key) or "").strip() else ("env" if resolved else "none")
        return {"name": name, "key": key, "model": _models.get(key) or default_model,
                "configured": bool(resolved), "api_key_masked": mask(resolved), "source": src,
                "enabled": llm_en.get(key, True), "endpoint": _llm_base(key) or endpoint}

    llm_providers = [
        _prov("Anthropic Claude", "anthropic", s.anthropic_model, "https://api.anthropic.com"),
        _prov("OpenAI GPT", "openai", s.openai_model, "https://api.openai.com"),
        _prov("Google Gemini", "gemini", s.gemini_model, "https://generativelanguage.googleapis.com"),
        _prov("Groq (open models)", "groq", s.groq_model, "https://api.groq.com/openai/v1"),
    ]
    market_data = [
        {"name": "NSE India", "key": "nse", "type": "public — no key required", "configured": True,
         "api_key_masked": "", "enabled": src_en.get("nse", True),
         "endpoints": ["https://www.nseindia.com/api/allIndices",
                       "https://www.nseindia.com/api/quote-equity?symbol={SYMBOL}",
                       "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv"]},
        {"name": "Yahoo Finance (fallback)", "key": "yahoo", "type": "public — no key required",
         "configured": True, "api_key_masked": "", "enabled": src_en.get("yahoo", True),
         "endpoints": ["https://query1.finance.yahoo.com/v8/finance/chart/{SYMBOL}.NS"]},
        {"name": "Zerodha Kite Connect", "key": "kite", "type": "licensed broker feed",
         "configured": bool(s.kite_api_key and s.kite_access_token),
         "api_key_masked": mask(s.kite_api_key), "enabled": src_en.get("kite", True),
         "endpoints": ["https://api.kite.trade"]},
        {"name": "Angel One SmartAPI", "key": "smartapi", "type": "licensed broker feed",
         "configured": bool(s.smartapi_key and s.smartapi_access_token),
         "api_key_masked": mask(s.smartapi_key), "enabled": src_en.get("smartapi", True),
         "endpoints": ["https://apiconnect.angelone.in"]},
        {"name": "Upstox", "key": "upstox", "type": "licensed broker feed",
         "configured": bool(s.upstox_access_token),
         "api_key_masked": mask(s.upstox_access_token), "enabled": src_en.get("upstox", True),
         "endpoints": ["https://api.upstox.com/v2"]},
    ]
    news_feeds = [{"name": n, "url": u} for n, u in FEEDS.items()]
    return {"llm_providers": llm_providers, "market_data": market_data,
            "news_feeds": news_feeds,
            "note": "Set a key here to store it in the database (masked) - it takes "
                    "priority over backend/.env. Leave the field blank to keep using the .env "
                    "value. Full keys never leave the server."}


# ── Broker-research RAG store ────────────────────────────────────
_MAX_RESEARCH_BYTES = 20 * 1024 * 1024  # 20 MB


@router.get("/research")
def research_list():
    from app.services import research
    return {"documents": research.list_documents(),
            "note": "Uploaded broker research grounds the AI assistant's answers "
                    "as cited reference material. The assistant reports it factually "
                    "and never presents it as buy/sell advice."}


@router.post("/research/upload")
async def research_upload(file: UploadFile = File(...), title: str = Form(""),
                          source: str = Form(""),
                          admin: User = Depends(require_admin)):
    """Upload a .pdf / .txt / .md research document into the RAG store."""
    from app.services import research
    data = await file.read()
    if len(data) > _MAX_RESEARCH_BYTES:
        raise HTTPException(413, "File too large (max 20 MB).")
    name = (file.filename or "").lower()
    if not name.endswith((".pdf", ".txt", ".md")):
        raise HTTPException(400, "Only .pdf, .txt or .md files are supported.")
    try:
        text = research.extract_text(file.filename, data)
        result = await research.ingest_document(
            title=title or file.filename, text=text, source=source,
            filename=file.filename, uploaded_by=admin.email)
    except ValueError as e:
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(500, f"Failed to ingest document: {e}")
    return result


class ResearchTextRequest(BaseModel):
    title: str = Field(min_length=1, max_length=300)
    text: str = Field(min_length=1)
    source: str = ""


@router.post("/research/text")
async def research_text(req: ResearchTextRequest, admin: User = Depends(require_admin)):
    """Paste research text directly (no file) into the RAG store."""
    from app.services import research
    try:
        return await research.ingest_document(
            title=req.title, text=req.text, source=req.source,
            filename="", uploaded_by=admin.email)
    except ValueError as e:
        raise HTTPException(422, str(e))


@router.delete("/research/{doc_id}")
def research_delete(doc_id: int, admin: User = Depends(require_admin)):
    from app.services import research
    if not research.delete_document(doc_id):
        raise HTTPException(404, "Document not found")
    audit_log("research_deleted_by", doc_id=doc_id, by=admin.email)
    return {"deleted": doc_id}


# ── LLM connectivity test (diagnose AI failures) ─────────────────
@router.post("/llm-test")
async def llm_test():
    """Probe each configured LLM provider with a 1-token call and report which
    work and which fail (and why)."""
    from app.config import get_settings
    from app.llm.providers import (AnthropicProvider, GeminiProvider, GroqProvider,
                                   OpenAIProvider)
    from app.services.app_settings import get_setting, llm_key
    s = get_settings()
    models = get_setting("llm_models") or {}
    candidates = [
        ("anthropic", AnthropicProvider, models.get("anthropic") or s.anthropic_model),
        ("openai", OpenAIProvider, models.get("openai") or s.openai_model),
        ("gemini", GeminiProvider, models.get("gemini") or s.gemini_model),
        ("groq", GroqProvider, models.get("groq") or s.groq_model),
    ]
    results = []
    for name, cls, model in candidates:
        if not llm_key(name):
            results.append({"provider": name, "model": model, "configured": False,
                            "ok": False, "detail": "no API key set (Admin -> Integrations or .env)"})
            continue
        try:
            p = cls(model)
            resp = await p.complete("You are a test.", "Reply with the single word OK.",
                                    max_tokens=64, temperature=0)   # >5 so reasoning models (gpt-oss) still emit content
            txt = (resp.text or "").strip()
            results.append({"provider": name, "model": model, "configured": True,
                            "ok": bool(txt),
                            "detail": txt[:40] if txt else "connected but returned empty output"})
        except Exception as e:
            results.append({"provider": name, "model": model, "configured": True,
                            "ok": False, "detail": f"{type(e).__name__}: {str(e)[:200]}"})
    any_ok = any(r["ok"] for r in results)
    audit_log("llm_test", any_ok=any_ok,
              results=[{"provider": r["provider"], "ok": r["ok"]} for r in results])
    return {"any_provider_working": any_ok, "results": results,
            "note": "If all show ok=false, the AI Assistant returns 'AI service "
                    "unavailable'. Fix the key/model in backend/.env and restart."}


# ── RBAC: roles & page access ────────────────────────────────────
@router.get("/pages")
def page_catalog():
    return {"pages": ALL_PAGES}


@router.get("/roles")
def list_roles():
    db = SessionLocal()
    try:
        return [{"id": r.id, "name": r.name, "pages": r.pages or [],
                 "is_admin": bool(r.is_admin),
                 "users": db.query(User).filter_by(role_id=r.id).count()}
                for r in db.query(Role).order_by(Role.name).all()]
    finally:
        db.close()


class RoleRequest(BaseModel):
    name: str = Field(min_length=1, max_length=60)
    pages: list[str] = []
    is_admin: bool = False


def _validate_pages(pages):
    bad = [p for p in pages if p not in ALL_PAGES]
    if bad:
        raise HTTPException(400, f"Unknown page(s): {', '.join(bad)}")


@router.post("/roles")
def create_role(req: RoleRequest, admin: User = Depends(require_admin)):
    _validate_pages(req.pages)
    db = SessionLocal()
    try:
        if db.query(Role).filter_by(name=req.name.strip()).first():
            raise HTTPException(409, "A role with that name already exists")
        role = Role(name=req.name.strip(), pages=req.pages, is_admin=req.is_admin)
        db.add(role); db.commit(); db.refresh(role)
        rid = role.id
    finally:
        db.close()
    audit_log("role_created", name=req.name, is_admin=req.is_admin, by=admin.email)
    return {"id": rid, "name": req.name}


@router.put("/roles/{role_id}")
def update_role(role_id: int, req: RoleRequest, admin: User = Depends(require_admin)):
    _validate_pages(req.pages)
    db = SessionLocal()
    try:
        role = db.get(Role, role_id)
        if not role:
            raise HTTPException(404, "Role not found")
        role.name = req.name.strip(); role.pages = req.pages; role.is_admin = req.is_admin
        db.commit()
    finally:
        db.close()
    audit_log("role_updated", id=role_id, by=admin.email)
    return {"id": role_id}


@router.delete("/roles/{role_id}")
def delete_role(role_id: int, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        role = db.get(Role, role_id)
        if not role:
            raise HTTPException(404, "Role not found")
        assigned = db.query(User).filter_by(role_id=role_id).count()
        if assigned:
            raise HTTPException(409, f"{assigned} user(s) still have this role; reassign first")
        db.delete(role); db.commit()
    finally:
        db.close()
    audit_log("role_deleted", id=role_id, by=admin.email)
    return {"deleted": role_id}


class UserRoleRequest(BaseModel):
    role_id: int | None = None


@router.patch("/users/{user_id}/role")
def set_user_role(user_id: int, req: UserRoleRequest, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        user = db.get(User, user_id)
        if not user:
            raise HTTPException(404, "User not found")
        if req.role_id is not None and not db.get(Role, req.role_id):
            raise HTTPException(404, "Role not found")
        user.role_id = req.role_id
        db.commit()
        result = {"id": user.id, "email": user.email, "role_id": user.role_id}
    finally:
        db.close()
    audit_log("user_role_set", **result, by=admin.email)
    return result


# ── Ops triggers (Agents screen) ─────────────────────────────────
@router.post("/refresh-news")
async def refresh_news_now(admin: User = Depends(require_admin)):
    from app.services import news_intel
    await news_intel.refresh_news()
    audit_log("news_refresh_manual", by=admin.email)
    return {"status": "news refreshed"}


# ── Chat audit (admin) ───────────────────────────────────────────
@router.get("/chat-audit")
def chat_audit(user_email: str = "", session: str = "", limit: int = 20, offset: int = 0):
    """Full conversation log: who asked what, when, the AI response, the LLM
    provider, confidence and latency. Filterable + paginated."""
    from app.db.database import ChatMessage
    db = SessionLocal()
    try:
        emails = {u.id: u.email for u in db.query(User).all()}
        q = db.query(ChatMessage)
        if user_email:
            uids = [uid for uid, em in emails.items()
                    if user_email.lower() in (em or "").lower()]
            q = q.filter(ChatMessage.user_id.in_(uids or [-1]))
        if session:
            q = q.filter(ChatMessage.session_id.like(f"%{session}%"))
        total = q.count()
        rows = (q.order_by(ChatMessage.created_at.desc())
                .offset(offset).limit(limit).all())
        out = [{
            "id": r.id, "time": str(r.created_at),
            "user": emails.get(r.user_id) or ("—" if r.user_id is None else f"user#{r.user_id}"),
            "session_id": r.session_id, "role": r.role, "content": r.content,
            "provider": (r.meta or {}).get("provider"),
            "confidence": (r.meta or {}).get("confidence"),
            "latency_ms": (r.meta or {}).get("latency_ms"),
            "sources": (r.meta or {}).get("n_sources"),
        } for r in rows]
        return {"total": total, "rows": out}
    finally:
        db.close()


# ── Branding (admin-uploaded logo / favicon) ─────────────────────
@router.post("/branding")
async def upload_branding(file: UploadFile = File(...), admin: User = Depends(require_admin)):
    """Upload a logo image (PNG/JPG/SVG/WebP). Stored as a data URI and used as
    the app logo and favicon everywhere."""
    import base64
    data = await file.read()
    if len(data) > 600 * 1024:
        raise HTTPException(413, "Logo too large (max 600 KB). Please upload a smaller image.")
    ct = (file.content_type or "").lower()
    if not ct.startswith("image/"):
        name = (file.filename or "").lower()
        ext = {"svg": "image/svg+xml", "png": "image/png", "jpg": "image/jpeg",
               "jpeg": "image/jpeg", "webp": "image/webp", "gif": "image/gif"}
        ct = next((v for k, v in ext.items() if name.endswith("." + k)), "")
        if not ct:
            raise HTTPException(400, "Upload a PNG, JPG, SVG, WebP or GIF image.")
    uri = f"data:{ct};base64," + base64.b64encode(data).decode()
    set_setting("brand_logo", uri)
    audit_log("branding_uploaded", by=admin.email, bytes=len(data), type=ct)
    return {"ok": True, "bytes": len(data)}


@router.delete("/branding")
def clear_branding(admin: User = Depends(require_admin)):
    set_setting("brand_logo", "")
    audit_log("branding_cleared", by=admin.email)
    return {"ok": True}


# ── App settings (DB-configurable) ───────────────────────────────
class SettingUpdate(BaseModel):
    key: str
    value: object


@router.get("/settings")
def get_app_settings():
    # Never expose raw API keys to the browser - blank them in the settings payload
    # (managed separately via /admin/llm-keys with masking).
    st = dict(all_settings())
    st["llm_api_keys"] = {}
    st["graph_client_secret"] = ""   # secret managed via /admin/email-config
    return {"settings": st, "defaults": DEFAULTS}


@router.put("/settings")
def update_setting(req: SettingUpdate, admin: User = Depends(require_admin)):
    try:
        set_setting(req.key, req.value)
    except (KeyError, ValueError) as e:
        raise HTTPException(400, str(e))
    note = "Saved."
    try:
        if req.key == "daily_scoring_hour":
            from app.main import reschedule_scoring
            ok = reschedule_scoring(req.value)
            note = (f"Scheduler updated live - next daily run at {int(req.value):02d}:00 IST."
                    if ok else "Saved, but live reschedule failed - restart to apply.")
        elif req.key == "news_refresh_minutes":
            from app.main import reschedule_news
            ok = reschedule_news(req.value)
            note = ("News refresh interval updated live."
                    if ok else "Saved, but live reschedule failed - restart to apply.")
    except Exception:
        note = "Saved - restart the backend to apply the new schedule."
    audit_log("setting_updated", key=req.key, value=req.value, by=admin.email)
    return {"key": req.key, "value": req.value, "note": note}


@router.get("/chat-feedback")
async def chat_feedback_list(rating: int = 0, limit: int = 50,
                             admin: User = Depends(require_admin)):
    """Assistant-quality view: thumbs up/down totals + recent rated answers
    (rating=-1 to see only negatives)."""
    db = SessionLocal()
    try:
        q = db.query(ChatFeedback)
        if rating in (1, -1):
            q = q.filter_by(rating=rating)
        rows = q.order_by(ChatFeedback.created_at.desc()).limit(min(max(limit, 1), 200)).all()
        up = db.query(ChatFeedback).filter_by(rating=1).count()
        down = db.query(ChatFeedback).filter_by(rating=-1).count()
        return {"up": up, "down": down, "items": [
            {"id": r.id, "rating": r.rating, "question": r.question, "answer": r.answer,
             "provider": r.provider, "session_id": r.session_id, "at": str(r.created_at)}
            for r in rows]}
    finally:
        db.close()


# ── Partner Open API key management (admin only) ─────────────────
from app.core.partner_auth import ALL_SCOPES, generate_key  # noqa: E402


class PartnerKeyCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=80)
    scopes: list[str] = Field(default_factory=lambda: list(ALL_SCOPES))
    rate_limit_per_min: int = Field(default=60, ge=1, le=6000)


def _key_public(k: PartnerKey) -> dict:
    return {"id": k.id, "name": k.name, "key_prefix": k.key_prefix,
            "scopes": k.scopes or [], "rate_limit_per_min": k.rate_limit_per_min,
            "is_active": k.is_active, "call_count": k.call_count or 0,
            "created_by": k.created_by or "", "created_at": fmt_ist(k.created_at),
            "last_used_at": fmt_ist(k.last_used_at)}


@router.get("/partner-keys")
def list_partner_keys():
    db = SessionLocal()
    try:
        rows = db.query(PartnerKey).order_by(PartnerKey.created_at.desc()).all()
        return {"keys": [_key_public(k) for k in rows], "available_scopes": list(ALL_SCOPES)}
    finally:
        db.close()


@router.post("/partner-keys")
def create_partner_key(req: PartnerKeyCreate, admin: User = Depends(require_admin)):
    bad = [s for s in req.scopes if s not in ALL_SCOPES]
    if bad:
        raise HTTPException(400, f"Unknown scope(s): {bad}. Allowed: {ALL_SCOPES}")
    if not req.scopes:
        raise HTTPException(400, "At least one scope is required")
    full, prefix, khash = generate_key()
    db = SessionLocal()
    try:
        row = PartnerKey(name=req.name.strip(), key_prefix=prefix, key_hash=khash,
                         scopes=req.scopes, rate_limit_per_min=req.rate_limit_per_min,
                         is_active=True, created_by=admin.email)
        db.add(row)
        db.commit()
        db.refresh(row)
        pub = _key_public(row)
    finally:
        db.close()
    audit_log("partner_key_created", name=req.name, by=admin.email, scopes=req.scopes)
    # The full key is returned ONCE here and never stored in plaintext.
    return {"api_key": full, "key": pub,
            "note": "Store this key now - it cannot be retrieved again."}


@router.post("/partner-keys/{key_id}/revoke")
def revoke_partner_key(key_id: int, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        row = db.get(PartnerKey, key_id)
        if not row:
            raise HTTPException(404, "Key not found")
        row.is_active = False
        db.commit()
    finally:
        db.close()
    audit_log("partner_key_revoked", key_id=key_id, by=admin.email)
    return {"ok": True}


@router.delete("/partner-keys/{key_id}")
def delete_partner_key(key_id: int, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        row = db.get(PartnerKey, key_id)
        if not row:
            raise HTTPException(404, "Key not found")
        db.delete(row)
        db.commit()
    finally:
        db.close()
    audit_log("partner_key_deleted", key_id=key_id, by=admin.email)
    return {"ok": True}


# ── LLM API key management (admin only) ──────────────────────────
def _mask_key(k: str) -> str:
    k = (k or "").strip()
    return ("\u2026" + k[-4:]) if len(k) >= 4 else ("set" if k else "")


class LlmKeyReq(BaseModel):
    provider: str
    key: str | None = None      # new key; "" clears the DB entry (falls back to .env)
    base: str | None = None     # optional base URL override


@router.get("/llm-keys")
def llm_keys_status():
    from app.config import get_settings
    s = get_settings()
    db_keys = all_settings().get("llm_api_keys") or {}
    bases = all_settings().get("llm_base_urls") or {}
    envmap = {"anthropic": s.anthropic_api_key, "openai": s.openai_api_key,
              "gemini": s.google_api_key, "groq": s.groq_api_key}
    out = []
    for p in ("anthropic", "openai", "gemini", "groq"):
        dbk = (db_keys.get(p) or "").strip()
        envk = (envmap.get(p) or "").strip()
        out.append({
            "provider": p,
            "source": "db" if dbk else ("env" if envk else "none"),
            "masked": _mask_key(dbk or envk),
            "base": bases.get(p) or "",
        })
    return {"keys": out}


@router.post("/llm-keys")
def set_llm_key(req: LlmKeyReq, admin: User = Depends(require_admin)):
    if req.provider not in ("anthropic", "openai", "gemini", "groq"):
        raise HTTPException(400, "unknown provider")
    if req.key is not None:
        keys = dict(all_settings().get("llm_api_keys") or {})
        kv = req.key.strip()
        if kv:
            keys[req.provider] = kv
        else:
            keys.pop(req.provider, None)     # cleared -> fall back to .env
        set_setting("llm_api_keys", keys)
    if req.base is not None:
        bases = dict(all_settings().get("llm_base_urls") or {})
        bv = req.base.strip()
        if bv:
            bases[req.provider] = bv
        else:
            bases.pop(req.provider, None)
        set_setting("llm_base_urls", bases)
    audit_log("llm_key_set", provider=req.provider, by=admin.email,
              has_key=bool((req.key or "").strip()))
    return {"ok": True}


# ── Invite-code management (admin) ───────────────────────────────
class InviteCreate(BaseModel):
    label: str = "admin"
    max_uses: int = 25
    count: int = 1


@router.get("/invite-codes")
def list_invite_codes():
    db = SessionLocal()
    try:
        rows = db.query(InviteCode).order_by(InviteCode.created_at.desc()).limit(500).all()
        owners = {u.id: u.email for u in db.query(User.id, User.email).all()}
        wl = db.query(Waitlist).count()
        return {"codes": [{
            "id": r.id, "code": r.code,
            "owner": owners.get(r.owner_user_id) if r.owner_user_id else "admin",
            "max_uses": r.max_uses, "used_count": r.used_count or 0,
            "is_active": r.is_active, "created_by": r.created_by,
            "created_at": fmt_ist(r.created_at)} for r in rows],
            "waitlist_count": wl}
    finally:
        db.close()


@router.post("/invite-codes")
def create_invite_codes(req: InviteCreate, admin: User = Depends(require_admin)):
    from app.core.registration import _gen_code
    n = max(1, min(int(req.count), 100))
    db = SessionLocal()
    created = []
    try:
        for _ in range(n):
            code = _gen_code(db)
            db.add(InviteCode(code=code, owner_user_id=None,
                              max_uses=max(1, int(req.max_uses)), used_count=0,
                              is_active=True, created_by=req.label or "admin"))
            created.append(code)
        db.commit()
    finally:
        db.close()
    audit_log("invite_codes_created", by=admin.email, count=n)
    return {"created": created}


@router.post("/invite-codes/{code_id}/toggle")
def toggle_invite_code(code_id: int, admin: User = Depends(require_admin)):
    db = SessionLocal()
    try:
        row = db.get(InviteCode, code_id)
        if not row:
            raise HTTPException(404, "Code not found")
        row.is_active = not row.is_active
        db.commit()
        return {"id": row.id, "is_active": row.is_active}
    finally:
        db.close()


@router.get("/user-activity")
def user_activity(from_: str = Query(None, alias="from"), to: str = Query(None)):
    """User-activity analytics for admins: totals & growth, acquisition breakdown,
    invite funnel, users table and waitlist. Date filter applies to new users /
    invites in the range; totals are all-time."""
    from collections import Counter
    from datetime import datetime, timedelta, timezone

    def aware(c):
        if c is None:
            return None
        return c.replace(tzinfo=timezone.utc) if c.tzinfo is None else c

    today = datetime.now(timezone.utc).date()
    try:
        d_to = datetime.strptime(to, "%Y-%m-%d").date() if to else today
    except Exception:
        d_to = today
    try:
        d_from = datetime.strptime(from_, "%Y-%m-%d").date() if from_ else (d_to - timedelta(days=29))
    except Exception:
        d_from = d_to - timedelta(days=29)
    start = datetime(d_from.year, d_from.month, d_from.day, tzinfo=timezone.utc)
    end = datetime(d_to.year, d_to.month, d_to.day, tzinfo=timezone.utc) + timedelta(days=1)

    db = SessionLocal()
    try:
        users = db.query(User).all()
        invs = db.query(Invitation).all()
        wl = db.query(Waitlist).order_by(Waitlist.id.desc()).all()
    finally:
        db.close()

    def in_range(c):
        c = aware(c)
        return bool(c and start <= c < end)

    total = len(users)
    verified = sum(1 for u in users if u.email_verified)
    admins = sum(1 for u in users if u.is_admin)
    new_users = [u for u in users if in_range(u.created_at)]

    buckets = {}
    d = d_from
    while d <= d_to:
        buckets[d.isoformat()] = 0
        d += timedelta(days=1)
    for u in new_users:
        k = aware(u.created_at).date().isoformat()
        if k in buckets:
            buckets[k] += 1
    growth = [{"date": k, "count": v} for k, v in buckets.items()]

    def acq(lst):
        google = sum(1 for u in lst if (u.auth_provider or "email") == "google")
        invited = sum(1 for u in lst if u.invited_by_code)
        ver = sum(1 for u in lst if u.email_verified)
        n = len(lst)
        return {"google": google, "email": n - google, "invited": invited, "self": n - invited,
                "verified": ver, "unverified": n - ver, "total": n}

    joined_emails = {u.email for u in users}
    sent = len(invs)
    delivered = sum(1 for i in invs if i.delivered)
    joined = sum(1 for i in invs if i.email in joined_emails)
    sent_in_range = sum(1 for i in invs if in_range(i.created_at))
    inviter_counts = Counter(i.inviter_user_id for i in invs)
    uid_name = {u.id: (u.full_name or u.email) for u in users}
    top_inviters = [{"user": uid_name.get(uid, str(uid)), "invites": n,
                     "joined": sum(1 for i in invs if i.inviter_user_id == uid and i.email in joined_emails)}
                    for uid, n in inviter_counts.most_common(10)]

    def row(u):
        src = "google" if (u.auth_provider or "email") == "google" else ("invited" if u.invited_by_code else "self")
        return {"id": u.id, "email": u.email, "full_name": u.full_name, "source": src,
                "auth_provider": u.auth_provider or "email", "invited_by_code": u.invited_by_code,
                "email_verified": bool(u.email_verified), "is_admin": bool(u.is_admin),
                "created_at": (aware(u.created_at).isoformat() if u.created_at else None),
                "signup_ip": u.signup_ip, "last_ip": u.last_ip,
                "last_login_at": (aware(u.last_login_at).isoformat() if u.last_login_at else None)}
    fallback = datetime.min.replace(tzinfo=timezone.utc)
    user_rows = [row(u) for u in sorted(users, key=lambda x: aware(x.created_at) or fallback, reverse=True)][:2000]

    return {
        "range": {"from": d_from.isoformat(), "to": d_to.isoformat()},
        "totals": {"users": total, "verified": verified, "unverified": total - verified,
                   "admins": admins, "new_in_range": len(new_users)},
        "growth": growth,
        "acquisition_range": acq(new_users),
        "acquisition_all": acq(users),
        "invites": {"sent": sent, "delivered": delivered, "joined": joined,
                    "sent_in_range": sent_in_range, "top_inviters": top_inviters},
        "waitlist": {"count": len(wl),
                     "list": [{"email": w.email, "created_at": (aware(w.created_at).isoformat() if w.created_at else None)} for w in wl[:2000]]},
        "users": user_rows,
    }


class WaitlistAction(BaseModel):
    email: EmailStr


@router.post("/waitlist/remove")
def waitlist_remove(req: WaitlistAction):
    db = SessionLocal()
    try:
        n = db.query(Waitlist).filter(Waitlist.email == req.email.lower()).delete()
        db.commit()
    finally:
        db.close()
    audit_log("waitlist_remove", email=req.email.lower())
    return {"removed": n}


@router.post("/waitlist/invite")
def waitlist_invite(req: WaitlistAction, admin: User = Depends(require_admin)):
    """Create a one-time invite code bound to the waitlisted email, email it, and
    remove them from the waitlist. Returns the code (share manually if no SMTP)."""
    from app.core import registration as reg
    from app.db.database import Invitation
    email = req.email.lower()
    db = SessionLocal()
    try:
        code = reg._gen_code(db)
        db.add(InviteCode(code=code, owner_user_id=admin.id, max_uses=1, used_count=0,
                          is_active=True, created_by="admin"))
        # Bind the code to this email (expires + can't be forwarded) and log it.
        if not db.query(Invitation.id).filter_by(inviter_user_id=admin.id, email=email).first():
            db.add(Invitation(inviter_user_id=admin.id, email=email, code=code,
                              status="sent", delivered=False))
        db.query(Waitlist).filter(Waitlist.email == email).delete()
        db.commit()
    finally:
        db.close()
    delivered = reg._send_invite_email(email, code, "The NIYTRI Team")
    audit_log("waitlist_invite", email=email, delivered=delivered)
    return {"ok": True, "code": code, "delivered": delivered}


@router.post("/waitlist/clear-all")
def waitlist_clear_all():
    db = SessionLocal()
    try:
        n = db.query(Waitlist).delete()
        db.commit()
    finally:
        db.close()
    audit_log("waitlist_clear_all", count=n)
    return {"removed": n}


class EmailConfigReq(BaseModel):
    provider: str | None = None
    graph_tenant_id: str | None = None
    graph_client_id: str | None = None
    graph_client_secret: str | None = None   # send blank to keep existing
    graph_sender: str | None = None
    support_email: str | None = None


@router.get("/email-config")
def email_config():
    from app.config import get_settings
    s = get_settings()
    a = all_settings()
    return {"provider": a.get("email_provider") or "smtp",
            "graph_tenant_id": a.get("graph_tenant_id") or "",
            "graph_client_id": a.get("graph_client_id") or "",
            "graph_sender": a.get("graph_sender") or "",
            "graph_secret_set": bool((a.get("graph_client_secret") or "").strip()),
            "smtp_configured": bool((s.smtp_host or "") and (s.smtp_from or "")),
            "smtp_from": s.smtp_from or "",
            "support_email": a.get("support_email") or ""}


@router.post("/email-config")
def set_email_config(req: EmailConfigReq, admin: User = Depends(require_admin)):
    try:
        if req.provider is not None:
            set_setting("email_provider", req.provider)
        for k in ("graph_tenant_id", "graph_client_id", "graph_sender"):
            v = getattr(req, k)
            if v is not None:
                set_setting(k, v.strip())
        if req.graph_client_secret is not None and req.graph_client_secret.strip():
            set_setting("graph_client_secret", req.graph_client_secret.strip())
        if req.support_email is not None and req.support_email.strip():
            set_setting("support_email", req.support_email.strip())
    except (KeyError, ValueError) as e:
        raise HTTPException(400, str(e))
    audit_log("email_config_update", provider=req.provider or get_setting("email_provider"))
    return {"ok": True}


class EmailTestReq(BaseModel):
    to: EmailStr


@router.post("/email-test")
def email_test(req: EmailTestReq, admin: User = Depends(require_admin)):
    from app.services import emailer
    plat = emailer.platform_name()
    delivered, err = emailer.send_email(
        str(req.to), f"{plat} - test email",
        f"This is a test email from your {plat} admin console.\n\n"
        "If you received it, outbound email is configured correctly.", kind="test")
    audit_log("email_test", to=str(req.to), delivered=delivered)
    return {"delivered": delivered, "error": err}


@router.get("/referral-tree")
def referral_tree():
    """Who invited whom: the referral graph + a leaderboard of top inviters.
    A member's referral = a user who joined using one of that member's codes."""
    db = SessionLocal()
    try:
        users = db.query(User).all()
        codes = db.query(InviteCode).all()
    finally:
        db.close()
    code_owner = {c.code: c.owner_user_id for c in codes}
    by_id = {u.id: u for u in users}
    inviter = {}
    for u in users:
        owner = code_owner.get(u.invited_by_code or "")
        inviter[u.id] = owner if owner in by_id else None
    children: dict[int, list[int]] = {}
    for uid, pid in inviter.items():
        if pid:
            children.setdefault(pid, []).append(uid)

    def network(uid, seen):
        total = 0
        for c in children.get(uid, []):
            if c in seen:
                continue
            seen.add(c)
            total += 1 + network(c, seen)
        return total

    nodes = []
    for u in users:
        direct = len(children.get(u.id, []))
        nodes.append({
            "id": u.id, "name": (u.full_name or u.email.split("@")[0]), "email": u.email,
            "inviter_id": inviter[u.id], "is_admin": bool(u.is_admin),
            "joined": u.created_at.isoformat() if u.created_at else None,
            "direct": direct, "network": network(u.id, set())})
    roots = [n["id"] for n in nodes if not n["inviter_id"]]
    leaderboard = sorted([n for n in nodes if n["direct"] > 0],
                         key=lambda x: (x["direct"], x["network"]), reverse=True)[:25]
    return {"nodes": nodes, "roots": roots, "leaderboard": leaderboard,
            "total_users": len(users)}


@router.get("/terms")
def admin_terms():
    a = all_settings()
    db = SessionLocal()
    try:
        versions = [{"seq": v.seq, "version": v.version, "created_by": v.created_by,
                     "created_at": fmt_ist(v.created_at)}
                    for v in db.query(TermsVersion).order_by(TermsVersion.seq.desc()).limit(50).all()]
    finally:
        db.close()
    return {"html": a.get("tos_html") or "", "version": a.get("tos_version") or "1.0",
            "seq": a.get("tos_seq") or 1, "min_seq": a.get("tos_min_seq") or 1,
            "support_email": a.get("support_email") or "", "versions": versions}


class TermsPublishReq(BaseModel):
    html: str
    version: str
    target: str = "new"   # new | existing | all


@router.post("/terms/publish")
def publish_terms(req: TermsPublishReq, admin: User = Depends(require_admin)):
    if req.target not in ("new", "existing", "all"):
        raise HTTPException(400, "target must be new, existing or all")
    new_seq = int(get_setting("tos_seq") or 1) + 1
    label = req.version.strip() or str(new_seq)
    set_setting("tos_html", req.html)
    set_setting("tos_version", label)
    set_setting("tos_seq", new_seq)
    if req.target in ("existing", "all"):
        set_setting("tos_min_seq", new_seq)   # forces everyone below to re-accept
    db = SessionLocal()
    try:
        db.add(TermsVersion(seq=new_seq, version=label, html=req.html, created_by=admin.email))
        db.commit()
    finally:
        db.close()
    audit_log("terms_publish", version=label, seq=new_seq, target=req.target, by=admin.email)
    return {"ok": True, "seq": new_seq, "version": label, "target": req.target}


@router.get("/terms/acceptances")
def terms_acceptances(limit: int = 300):
    db = SessionLocal()
    try:
        rows = db.query(TosAcceptance).order_by(TosAcceptance.accepted_at.desc()).limit(limit).all()
        return {"acceptances": [{"email": r.email, "version": r.version, "seq": r.seq,
                                 "ip": r.ip, "accepted_at": fmt_ist(r.accepted_at)} for r in rows]}
    finally:
        db.close()


@router.get("/email-logs")
def email_logs(kind: str = "", limit: int = 300):
    db = SessionLocal()
    try:
        q = db.query(EmailLog)
        if kind:
            q = q.filter(EmailLog.kind == kind)
        rows = q.order_by(EmailLog.created_at.desc()).limit(limit).all()
        kinds = sorted({k[0] for k in db.query(EmailLog.kind).distinct().all() if k[0]})
        return {"kinds": kinds, "logs": [{"to": r.to_addr, "subject": r.subject, "kind": r.kind,
                "provider": r.provider, "delivered": r.delivered, "error": r.error,
                "created_at": fmt_ist(r.created_at)} for r in rows]}
    finally:
        db.close()


# ── Price data — EOD OHLCV backfill / daily refresh / coverage ──────────────
@router.get("/prices/summary")
def prices_summary():
    """Coverage stats + current load progress for the Admin price panel."""
    from app.services import prices
    return {**prices.summary(), "status": prices.STATE}


@router.post("/prices/backfill")
def prices_backfill(years: int = 3):
    """Start a full (resumable) history backfill in the background."""
    from app.services import prices
    if not prices.start_backfill_bg(years=max(1, min(int(years), 10))):
        raise HTTPException(409, "A price load is already running.")
    audit_log("prices_backfill_started", years=years)
    return {"started": True, "years": years}


@router.post("/prices/daily-now")
def prices_daily_now():
    """Run the daily incremental refresh now (last ~month per symbol)."""
    import asyncio
    import threading
    from app.services import prices
    if prices.STATE.get("running"):
        raise HTTPException(409, "A price load is already running.")
    threading.Thread(target=lambda: asyncio.run(prices.daily_update()), daemon=True).start()
    audit_log("prices_daily_now")
    return {"started": True}


# ── Index master — refresh NIFTY 50 / 500 constituents (then daily scoring uses it) ─
@router.post("/universe/refresh")
def universe_refresh():
    """Refresh the NIFTY 50 / 500 membership in the instruments master (NSE with
    fallback to the bundled lists). Runs automatically before daily scoring too."""
    from app.services import universe
    res = universe.refresh_universe()
    audit_log("universe_refresh", result=res)
    return res
